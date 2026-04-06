from __future__ import annotations

import hashlib
from collections import defaultdict, deque
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Callable

import pandas as pd
import streamlit as st
from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
from openpyxl import load_workbook  # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import PatternFill  # pyright: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter  # pyright: ignore[reportMissingModuleSource]


@dataclass(frozen=True)
class MergedRow:
    values: tuple[Any, ...]
    status: str
    changed_columns: frozenset[int]


@dataclass(frozen=True)
class RowRecord:
    row: int
    values: tuple[Any, ...]
    row_hash: str


def is_csv_file(file_name: str) -> bool:
    return file_name.lower().endswith(".csv")


@st.cache_data(show_spinner=False)
def get_sheet_names(file_bytes: bytes, file_name: str) -> list[str]:
    if is_csv_file(file_name):
        return ["CSV"]

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def load_tabular_rows(file_bytes: bytes, file_name: str, sheet_name: str) -> list[tuple[Any, ...]]:
    if is_csv_file(file_name):
        dataframe = pd.read_csv(BytesIO(file_bytes), header=None)
        rows: list[tuple[Any, ...]] = []
        for row in dataframe.itertuples(index=False, name=None):
            normalized = tuple(None if pd.isna(value) else value for value in row)
            rows.append(normalized)
        return rows

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        return [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def compute_row_hash(row: tuple) -> str:
    """Compute SHA256 hash of row content for comparison."""
    row_str = "|".join("" if cell is None else str(cell) for cell in row)
    return hashlib.sha256(row_str.encode()).hexdigest()


def get_changed_columns(row_1: tuple[Any, ...], row_2: tuple[Any, ...], max_columns: int) -> frozenset[int]:
    changed_columns = {
        column_index
        for column_index in range(1, max_columns + 1)
        if (row_1[column_index - 1] if column_index <= len(row_1) else None)
        != (row_2[column_index - 1] if column_index <= len(row_2) else None)
    }
    return frozenset(changed_columns)


def count_matching_cells(row_1: tuple[Any, ...], row_2: tuple[Any, ...], max_columns: int) -> int:
    return sum(
        1
        for column_index in range(1, max_columns + 1)
        if (row_1[column_index - 1] if column_index <= len(row_1) else None)
        == (row_2[column_index - 1] if column_index <= len(row_2) else None)
    )


def match_updated_rows(
    unmatched_file1: list[RowRecord],
    unmatched_file2: list[RowRecord],
    max_columns: int,
) -> tuple[dict[int, RowRecord], set[int]]:
    if not unmatched_file1 or not unmatched_file2:
        return {}, set()

    min_match_score = max(1, int(max_columns * 0.7))
    candidates: list[tuple[int, int, int, RowRecord, RowRecord]] = []

    for row_1 in unmatched_file1:
        for row_2 in unmatched_file2:
            score = count_matching_cells(row_1.values, row_2.values, max_columns)
            if score >= min_match_score:
                # Secondary sort key prefers rows closer in original position.
                row_distance = abs(row_1.row - row_2.row)
                candidates.append((score, -row_distance, row_1.row, row_1, row_2))

    candidates.sort(reverse=True)

    pair_map: dict[int, RowRecord] = {}
    consumed_file2_rows: set[int] = set()
    consumed_file1_rows: set[int] = set()

    for _, _, row_1_number, row_1, row_2 in candidates:
        if row_1_number in consumed_file1_rows or row_2.row in consumed_file2_rows:
            continue
        pair_map[row_1_number] = row_2
        consumed_file1_rows.add(row_1_number)
        consumed_file2_rows.add(row_2.row)

    return pair_map, consumed_file2_rows


def load_sheet_rows(raw_rows: list[tuple[Any, ...]], max_columns: int) -> list[RowRecord]:
    row_records: list[RowRecord] = []
    for row_number, values in enumerate(raw_rows, start=1):
        padded_values = tuple(
            values[column_index] if column_index < len(values) else None
            for column_index in range(max_columns)
        )
        row_records.append(RowRecord(row=row_number, values=padded_values, row_hash=compute_row_hash(padded_values)))
    return row_records


def build_merged_rows(
    rows_1_data: list[RowRecord],
    rows_2_data: list[RowRecord],
    max_columns: int,
    progress_callback: Callable[[int, int], None] | None = None,
) -> tuple[list[MergedRow], dict[str, int]]:
    matched_file2: set[int] = set()
    file2_hash_to_indices: dict[str, deque[int]] = defaultdict(deque)

    for index, row_record in enumerate(rows_2_data):
        file2_hash_to_indices[row_record.row_hash].append(index)

    exact_pair_map: dict[int, int] = {}
    unmatched_file1: list[RowRecord] = []

    for current_index, row_record in enumerate(rows_1_data, start=1):
        if progress_callback and rows_1_data:
            progress_callback(current_index, len(rows_1_data))

        candidate_indices = file2_hash_to_indices.get(row_record.row_hash)
        if candidate_indices:
            matched_index = candidate_indices.popleft()
            matched_file2.add(matched_index)
            exact_pair_map[row_record.row] = rows_2_data[matched_index].row
            continue

        unmatched_file1.append(row_record)

    unmatched_file2 = [row_record for index, row_record in enumerate(rows_2_data) if index not in matched_file2]

    pair_map, consumed_file2_rows = match_updated_rows(unmatched_file1, unmatched_file2, max_columns)
    paired_rows = len(pair_map)
    exact_matched_file1_rows = set(exact_pair_map)
    all_consumed_file2_rows = set(exact_pair_map.values()) | consumed_file2_rows

    file2_by_row: dict[int, RowRecord] = {row_record.row: row_record for row_record in rows_2_data}
    inserted_file2_rows = {row_record.row for row_record in rows_2_data if row_record.row not in all_consumed_file2_rows}
    emitted_insert_rows: set[int] = set()

    merged_rows: list[MergedRow] = []
    next_file2_row = 1

    def emit_inserts_before(target_file2_row: int) -> None:
        nonlocal next_file2_row
        while next_file2_row < target_file2_row and next_file2_row <= len(rows_2_data):
            if next_file2_row in inserted_file2_rows:
                insert_row = file2_by_row[next_file2_row]
                merged_rows.append(
                    MergedRow(
                        values=insert_row.values,
                        status="inserted",
                        changed_columns=frozenset(range(1, max_columns + 1)),
                    )
                )
                emitted_insert_rows.add(next_file2_row)
            next_file2_row += 1

    for row_record in rows_1_data:
        exact_file2_row = exact_pair_map.get(row_record.row)
        if exact_file2_row is not None:
            emit_inserts_before(exact_file2_row)
            merged_rows.append(MergedRow(values=row_record.values, status="unchanged", changed_columns=frozenset()))
            if next_file2_row <= exact_file2_row:
                next_file2_row = exact_file2_row + 1
            continue

        matched_row_2 = pair_map.get(row_record.row)
        if matched_row_2 is not None:
            emit_inserts_before(matched_row_2.row)
            changed_columns = get_changed_columns(row_record.values, matched_row_2.values, max_columns)
            merged_rows.append(
                MergedRow(values=row_record.values, status="updated_file1", changed_columns=changed_columns)
            )
            merged_rows.append(
                MergedRow(values=matched_row_2.values, status="updated_file2", changed_columns=changed_columns)
            )
            if next_file2_row <= matched_row_2.row:
                next_file2_row = matched_row_2.row + 1
            continue

        merged_rows.append(
            MergedRow(values=row_record.values, status="deleted", changed_columns=frozenset(range(1, max_columns + 1)))
        )

    for file2_row in range(next_file2_row, len(rows_2_data) + 1):
        if file2_row in inserted_file2_rows and file2_row not in emitted_insert_rows:
            insert_row = file2_by_row[file2_row]
            merged_rows.append(
                MergedRow(values=insert_row.values, status="inserted", changed_columns=frozenset(range(1, max_columns + 1)))
            )

    # Safety: in rare non-monotonic matches, ensure no insert row is dropped.
    for file2_row in sorted(inserted_file2_rows - emitted_insert_rows):
        if file2_row < next_file2_row:
            insert_row = file2_by_row[file2_row]
            merged_rows.append(
                MergedRow(values=insert_row.values, status="inserted", changed_columns=frozenset(range(1, max_columns + 1)))
            )

    stats = {
        "rows_compared": max(len(rows_1_data), len(rows_2_data)),
        "columns_compared": max_columns,
        "unchanged_rows_count": len(exact_pair_map),
        "updated_rows_count": paired_rows,
        "deleted_rows_count": len(unmatched_file1) - paired_rows,
        "added_rows_count": len(unmatched_file2) - paired_rows,
        "merged_rows_count": len(merged_rows),
    }
    return merged_rows, stats


def create_merged_workbook(merged_rows: list[MergedRow], max_columns: int) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Merged Comparison"

    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    green_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

    for row_index, merged_row in enumerate(merged_rows, start=1):
        for column_index in range(1, max_columns + 1):
            value = merged_row.values[column_index - 1] if column_index <= len(merged_row.values) else None
            cell = worksheet.cell(row=row_index, column=column_index, value=value)

            if merged_row.status == "deleted":
                cell.fill = red_fill
            elif merged_row.status == "inserted":
                cell.fill = green_fill
            elif merged_row.status == "updated_file1" and column_index in merged_row.changed_columns:
                cell.fill = red_fill
            elif merged_row.status == "updated_file2" and column_index in merged_row.changed_columns:
                cell.fill = green_fill

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    return buffer.getvalue()


def build_preview_frame(merged_rows: list[MergedRow], max_columns: int) -> pd.DataFrame:
    preview_rows: list[list[Any]] = []
    for merged_row in merged_rows:
        row_values = [
            merged_row.values[column_index - 1] if column_index <= len(merged_row.values) else None
            for column_index in range(1, max_columns + 1)
        ]
        preview_rows.append(row_values)

    column_names = [get_column_letter(col) for col in range(1, max_columns + 1)]
    return pd.DataFrame(preview_rows, columns=column_names)


def style_preview_frame(preview_frame: pd.DataFrame, merged_rows: list[MergedRow]) -> pd.io.formats.style.Styler:
    red_hex = "#f4cccc"
    green_hex = "#d9ead3"

    def style_row(row: pd.Series) -> list[str]:
        merged_row = merged_rows[row.name]
        styles: list[str] = []

        for column_index in range(1, len(row) + 1):
            if merged_row.status == "deleted":
                styles.append(f"background-color: {red_hex}")
            elif merged_row.status == "inserted":
                styles.append(f"background-color: {green_hex}")
            elif merged_row.status == "updated_file1" and column_index in merged_row.changed_columns:
                styles.append(f"background-color: {red_hex}")
            elif merged_row.status == "updated_file2" and column_index in merged_row.changed_columns:
                styles.append(f"background-color: {green_hex}")
            else:
                styles.append("")
        return styles

    return preview_frame.style.apply(style_row, axis=1)


def compare_workbooks(
    file_1_bytes: bytes,
    file_2_bytes: bytes,
    file_1_name: str,
    file_2_name: str,
    sheet_1: str,
    sheet_2: str,
    progress_callback: Callable[[int, int], None] | None = None,
) -> tuple[list[MergedRow], dict[str, int], int]:
    rows_1_raw = load_tabular_rows(file_1_bytes, file_1_name, sheet_1)
    rows_2_raw = load_tabular_rows(file_2_bytes, file_2_name, sheet_2)

    max_columns = max(
        max((len(row) for row in rows_1_raw), default=0),
        max((len(row) for row in rows_2_raw), default=0),
    )

    rows_1_data = load_sheet_rows(rows_1_raw, max_columns)
    rows_2_data = load_sheet_rows(rows_2_raw, max_columns)

    merged_rows, stats = build_merged_rows(rows_1_data, rows_2_data, max_columns, progress_callback)
    return merged_rows, stats, max_columns


def render_results(
    merged_rows: list[MergedRow],
    stats: dict[str, int],
    max_columns: int,
) -> None:
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Rows", stats["rows_compared"])
    col2.metric("Columns", stats["columns_compared"])
    col3.metric("Merged Rows", stats["merged_rows_count"])
    col4.metric("Unchanged", stats["unchanged_rows_count"])
    col5.metric("Changes", stats["updated_rows_count"] + stats["deleted_rows_count"] + stats["added_rows_count"])

    st.caption("Legend: deleted rows are red, inserted rows are green, and updated rows appear twice with only the changed cells highlighted.")

    merged_workbook_bytes = create_merged_workbook(merged_rows, max_columns)
    st.download_button(
        "Download merged workbook",
        data=merged_workbook_bytes,
        file_name="excel_merged_comparison.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    show_preview = st.toggle("Preview merged output in app", value=True)
    if show_preview:
        st.subheader("Preview")
        preview_frame = build_preview_frame(merged_rows, max_columns)
        styled_preview = style_preview_frame(preview_frame, merged_rows)
        st.dataframe(styled_preview, use_container_width=True, hide_index=True)

    if stats["updated_rows_count"] == 0 and stats["deleted_rows_count"] == 0 and stats["added_rows_count"] == 0:
        st.success("No differences found in the compared sheets.")


st.set_page_config(page_title="Excel Cell Diff", page_icon="📊", layout="wide")

st.title("Excel Cell Difference Finder")
st.write("Upload two Excel/CSV files, pick sheets when needed, and compare row values.")
st.caption("Supports `.xlsx`, `.xlsm`, and `.csv` files.")

left, right = st.columns(2)
with left:
    file_1 = st.file_uploader("First file", type=["xlsx", "xlsm", "csv"])
with right:
    file_2 = st.file_uploader("Second file", type=["xlsx", "xlsm", "csv"])

if file_1 and file_2:
    file_1_bytes = file_1.getvalue()
    file_2_bytes = file_2.getvalue()

    try:
        sheet_names_1 = get_sheet_names(file_1_bytes, file_1.name)
        sheet_names_2 = get_sheet_names(file_2_bytes, file_2.name)
    except Exception as exc:
        st.error(f"Could not read one of the files: {exc}")
        st.stop()

    if len(sheet_names_1) == 1:
        sheet_1 = sheet_names_1[0]
        st.caption(f"First file source: {sheet_1}")
    else:
        sheet_1 = st.selectbox("Sheet from first file", sheet_names_1)

    if len(sheet_names_2) == 1:
        sheet_2 = sheet_names_2[0]
        st.caption(f"Second file source: {sheet_2}")
    else:
        sheet_2 = st.selectbox("Sheet from second file", sheet_names_2)

    compare_clicked = st.button("Compare files", type="primary")

    if compare_clicked:
        progress = st.progress(0)
        status = st.empty()

        def update_progress(current_row: int, total_rows: int) -> None:
            progress_value = int((current_row / total_rows) * 100) if total_rows else 100
            progress.progress(min(progress_value, 100))
            status.write(f"Comparing row {current_row} of {total_rows}...")

        with st.spinner("Comparing workbook cells..."):
            merged_rows, stats, max_columns = compare_workbooks(
                file_1_bytes=file_1_bytes,
                file_2_bytes=file_2_bytes,
                file_1_name=file_1.name,
                file_2_name=file_2.name,
                sheet_1=sheet_1,
                sheet_2=sheet_2,
                progress_callback=update_progress,
            )

        progress.progress(100)
        status.write("Comparison complete.")
        render_results(merged_rows, stats, max_columns)
else:
    st.info("Upload both files to begin.")
